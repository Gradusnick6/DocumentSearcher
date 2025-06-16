import openpyxl
from openpyxl.styles import Alignment
import os
from Model.DirectoryWork import SystemDirectoryHandler


class ExcelHandler:
    def __init__(self, file_path):
        """
        Создает файл при вызове конструктора
        :param file_path: str
        """
        try:
            self.__file_path = SystemDirectoryHandler.get_unique_name_for_exist_file(file_path)
            self.__workbook = self.create_file()
        except Exception as ex:
            raise ex

    def create_file(self):
        """
        :return: Workbook
        """
        try:
            if os.path.exists(self.__file_path):
                os.remove(self.__file_path)
            workbook = openpyxl.Workbook()
            workbook.save(self.__file_path)
            return workbook
        except Exception as ex:
            raise ex

    def save(self):
        """
        Исключения ValueError()
        :return: None
        """
        try:
            if self.__workbook:
                self.__workbook.save(self.__file_path)
            else:
                raise ValueError()
        except Exception as ex:
            raise ex

    def delete_sheet(self, sheet_name):
        """
        :param sheet_name: str
        :return: None
        """
        try:
            if sheet_name in self.__workbook.sheetnames:
                del self.__workbook[sheet_name]
        except Exception as ex:
            raise ex

    def create_sheet(self, sheet_name):
        """
        :param sheet_name: str
        :return: Sheet. объект созданного листа
        """
        try:
            if sheet_name not in self.__workbook.sheetnames:
                self.__workbook.create_sheet(sheet_name)
            return self.__workbook[sheet_name]
        except Exception as ex:
            raise ex

    def format_columns(self):
        """
        форматируем ширину колонки по самому длинному значению таблицы
        :return: None
        """
        try:
            for sheet in self.__workbook:
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Получаем букву колонки
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2  # Добавляем немного пространства для удобства
                    sheet.column_dimensions[column].width = adjusted_width
            self.save()
        except Exception as ex:
            raise ex

    def set_align(self):
        """
        настраиваем выравнивание текста в ячейках на "center" по горизонтали и вертикали
        :return: None
        """
        try:
            alignment = Alignment(horizontal='center', vertical='center')
            for sheet in self.__workbook:
                # Настраиваем выравнивание для всех ячеек на листе
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = alignment
            self.save()
        except Exception as ex:
            raise ex
